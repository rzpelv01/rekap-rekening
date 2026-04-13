#!/usr/bin/env python3
"""
BRI Rekening Koran → Excel Rekap
Cara pakai:
  python bri_rekap.py file.pdf
  python bri_rekap.py folder_berisi_pdf/
"""

import sys, re, os
from collections import defaultdict
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("Install dulu: pip install pdfplumber openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Install dulu: pip install openpyxl")
    sys.exit(1)

MONTHS_ID = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]

# Keyword yang PASTI bukan penjualan meski kredit
# Keyword PASTI bukan penjualan
NON_PENJ_KW = [
    # Bunga & biaya bank — semua format
    "INTEREST ON",
    "INTEREST ON ACCOUNT",
    "BUNGA TABUNGAN",
    "BUNGA DEPOSITO",
    "BUNGA GIRO",
    "JASA GIRO",        # BNI: "JASA GIRO/BUNGA"
    "BUNGA ",           # "Bunga 14201" di Mandiri
    " TAX",
    "PAJAK ",           # "Pajak 14201" di Mandiri
    "PPH",              # BNI: pajak bunga
    "BIAYA ADM",
    "BIAYA PROVISI",
    "BIAYA ADMINISTRASI",
    "ADMINISTRASI",
    "FEE SLCRD",        # biaya kartu BRI
    "FEE SLCRED",
    # ATS — auto sweep/top-up antar rekening (semua format, setelah .upper())
    "::ATS",         # ESB:T:0172167:S::ats → upper → ::ATS
    "NAS:RELOAD",
    "NAS:BILL",
    # Transfer internal
    "IFT TO",
    "FROM:0",
    "TO:0",
    "OVERBOOKING",
    "BY SURAT REF",
    "PMDH BUKUAN",
    "OB ESCROW",
    "ESCROW",
    "ESB:RTGS:",
    # Angsuran / pinjaman bank
    "LN INST.",
    # Pencairan kredit
    "PENCAIRAN",
    "CAIRKAN",
    "CAIR KRD",
    "PENCAIRAN KRD",
    "PMBY PINJ",
    "PLAFON",
    "FASILITAS KRD",
    "KMK",
    "KPR",
    "DROPING",
    "DROPPING",
    "REALISASI KRD",
    # Rekening pinjaman BRI — pembayaran & administrasi
    "*BAYAR POKOK",          # pembayaran pokok pinjaman
    "BAYAR POKOK",
    "*PEMBEBANAN BUNGA",     # pembebanan bunga pinjaman
    "PEMBEBANAN BUNGA",
    "PRINCIPAL PAYMENT",     # principal payment split out
    "LOCAL CHEQUE REPAYMENT",
    "PINDAH BUKU",           # pemindahbukuan internal
    "PNJMN ",                # pemberian pinjaman (PNJMN PANGAN LESTARI)
]

# Whole-word regex (hati-hati false positive)
NON_PENJ_WHOLE = [
    r'\bNAS\b',      # kode NAS sebagai kata utuh (bukan NASIONAL, ANAS, dll)
    r'\bBRIVA\b',
]

# Keyword Penjualan eksplisit
PENJ_KW = [
    # Pemerintah / APBD / APBN
    "SP2D", "SPAN", "KASDA", "APBD", "APBN",
    "PEMKOT", "PEMKAB", "PEMPROV", "PEMERINTAH",
    # Kesehatan
    "RSUD", "RSIA", "RSUP", "RSU ", "RS ",
    "RUMAH SAKIT", "PUSKESMAS", "KLINIK", "LABORATORIUM",
    # Pendidikan
    "SEKOLAH", "UNIVERSITAS", "UNIV ", "AKADEMI", "MADRASAH",
    "PESANTREN", "SDN ", "SMPN ", "SMAN ",
    # BUMN / perusahaan besar
    "PLN ", "PDAM", "PERTAMINA", "TELKOM",
    "PUPUK KALTIM", "PUPUK INDONESIA", "INKA MULTI",
    # Pola pembayaran
    "PAYMENT PT", "PEMBAYARAN PT", "BANK MANDIRI-PEMBAYARA",
    "BANK BNI-PEMB", "S2P",
    # Kode internal diketahui = penjualan
    "ESB:INDS:",
]

# Kata yang menandai akhir blok transaksi (baris summary / footer)
STOP_KW = [
    'Saldo Awal','Opening Balance','Terbilang','In Words',
    'Biaya materai','Revenue Stamp','Apabila terdapat','In the case',
    'Salinan rekening','The copy of','Apabila ada perubahan','Should there be',
    'Created By','BRISIM',
]

# Baris header tabel yang harus dilewati
SKIP_KW = [
    'Tanggal Transaksi','Transaction Date','LAPORAN TRANSAKSI',
    'Halaman','Page ','Printed By','Business Unit',
    'Nama Produk','Product Name','Valuta','Currency',
    'No. Rekening','Account No','Kepada Yth','Transaction Period',
    'Transaction Periode','STATEMENT OF','Statement Date',
    'Tanggal Laporan',
]

NUM_RE    = re.compile(r'^[\d,]+\.\d{2}$')
DATE_RE   = re.compile(r'^\d{2}/\d{2}/\d{2}$')
TIME_RE   = re.compile(r'^\d{2}:\d{2}:\d{2}$')
TELLER_RE = re.compile(r'^\d{7}$')

# ── Warna Excel ────────────────────────────────────────────────────────────────
CLR = {
    "title_bg" : "1F3864",  "title_fg" : "FFFFFF",
    "sub_bg"   : "2E75B6",  "sub_fg"   : "FFFFFF",
    "hdr_bg"   : "2E75B6",  "hdr_fg"   : "FFFFFF",
    "total_bg" : "FFF2CC",
    "alt_bg"   : "F2F7FB",
    "penj_bg"  : "E2EFDA",
    "debet_fg" : "C00000",
    "kredit_fg": "375623",
    "border"   : "BDD7EE",
}
NUM_FMT = '#,##0.00'

def thin_border():
    s = Side(style='thin', color=CLR["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def af(clr): return PatternFill("solid", fgColor=clr)

def style_hdr(ws, row_num, bg=None, fg=None):
    bg = bg or CLR["hdr_bg"]; fg = fg or CLR["hdr_fg"]
    for cell in ws[row_num]:
        cell.font      = Font(name='Arial', bold=True, color=fg, size=10)
        cell.fill      = af(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border()

def style_total(ws, row_num):
    for cell in ws[row_num]:
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.fill = af(CLR["total_bg"]); cell.border = thin_border()

def reg(bold=False, color="000000"):
    return Font(name='Arial', bold=bold, color=color, size=10)

# ── Deteksi kolom dari baris header tabel ─────────────────────────────────────
def detect_columns(rows):
    """
    Cari baris header 'Tanggal Transaksi ... Debet ... Kredit ... Saldo'
    dan gunakan posisi X-nya untuk menentukan batas kolom.
    Fallback ke nilai default jika tidak ditemukan.
    """
    for y in sorted(rows.keys()):
        row_words = sorted(rows[y], key=lambda w: w['x0'])
        texts = [w['text'] for w in row_words]
        line  = ' '.join(texts)
        if 'Debet' in line and 'Kredit' in line and 'Saldo' in line:
            col_map = {}
            for w in row_words:
                t = w['text']
                if t == 'Debet':  col_map['debet']  = w['x0']
                if t == 'Kredit': col_map['kredit'] = w['x0']
                if t == 'Saldo':  col_map['saldo']  = w['x0']
                if t in ('Teller','User'): col_map['teller'] = w['x0']
            if 'debet' in col_map:
                # Teller biasanya ~30px sebelum Debet
                teller_x = col_map.get('teller', col_map['debet'] - 30)
                return {
                    'teller': teller_x - 5,
                    'debet' : col_map['debet']  - 15,
                    'kredit': col_map['kredit'] - 15,
                    'saldo' : col_map['saldo']  - 15,
                }
    # Default (format lama)
    return {'teller': 295, 'debet': 360, 'kredit': 465, 'saldo': 560}

# ── Temukan Y baris summary (Saldo Awal) ──────────────────────────────────────
def find_summary_y(rows):
    for y in sorted(rows.keys()):
        line = ' '.join(w['text'] for w in sorted(rows[y], key=lambda w: w['x0']))
        if 'Saldo Awal' in line or 'Opening Balance' in line:
            return y
    return 99999

# ── Parse satu PDF ─────────────────────────────────────────────────────────────
def _is_bca_pdf(pdf_path):
    """Deteksi apakah PDF adalah format BCA (bukan BRI)."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            words = pdf.pages[0].extract_words()
            text = ' '.join(w['text'] for w in words)
            return ('REKENING GIRO' in text or 'REKENING TABUNGAN' in text) and \
                   'PERIODE' in text and 'MUTASI' in text and 'TANGGAL' in text
    except:
        return False


def _parse_pdf_bca(pdf_path):
    """Parser khusus untuk format e-statement BCA."""
    import re as _re

    meta = {
        "accountNo": "", "companyName": "", "period": "",
        "opening": 0, "totalDebet": 0, "totalKredit": 0, "closing": 0,
        "currency": "IDR"
    }
    transactions = []
    year = ""

    NUM_RE_BCA  = re.compile(r'^[\d,]+\.\d{2}$')
    DATE_RE_BCA = re.compile(r'^\d{2}/\d{2}$')

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            rows = {}
            for w in words:
                y = round(w['top'] / 2) * 2
                rows.setdefault(y, []).append(w)

            # ── Meta dari halaman 1 ──
            if page_num == 0:
                for y in sorted(rows.keys()):
                    rw   = sorted(rows[y], key=lambda w: w['x0'])
                    line = ' '.join(w['text'] for w in rw)

                    if not meta["accountNo"]:
                        m = re.search(r'NO\.\s*REKENING\s*:\s*(\d{10,})', line)
                        if m: meta["accountNo"] = m.group(1)

                    if not meta["companyName"] and y > 70 and y < 100:
                        left = [w['text'] for w in rw if w['x0'] < 200
                                and w['text'] not in ('PT','CV','KCU','KCP')]
                        if left:
                            candidate = ' '.join(left).strip()
                            if len(candidate) >= 3:
                                meta["companyName"] = candidate

                    if not year:
                        m = re.search(r'PERIODE\s*:\s*(\w+)\s+(\d{4})', line)
                        if m:
                            bulan_map = {
                                'JANUARI':'Jan','FEBRUARI':'Feb','MARET':'Mar',
                                'APRIL':'Apr','MEI':'Mei','JUNI':'Jun',
                                'JULI':'Jul','AGUSTUS':'Agu','SEPTEMBER':'Sep',
                                'OKTOBER':'Okt','NOVEMBER':'Nov','DESEMBER':'Des'
                            }
                            year = m.group(2)
                            meta["period"] = f"01/01/{year[2:]} - 31/12/{year[2:]}"

                    if 'MATA UANG' in line or 'CURRENCY' in line:
                        cm = re.search(r'\b(USD|EUR|SGD|IDR|CNY|JPY)\b', line)
                        if cm: meta["currency"] = cm.group(1)

            # ── Summary di halaman terakhir ──
            for y in sorted(rows.keys()):
                rw   = sorted(rows[y], key=lambda w: w['x0'])
                line = ' '.join(w['text'] for w in rw)
                m = re.search(r'SALDO\s+AWAL\s*:\s*([\d,]+\.\d{2})', line)
                if m and not meta["opening"]:
                    meta["opening"] = float(m.group(1).replace(',',''))
                m = re.search(r'SALDO\s+AKHIR\s*:\s*([\d,]+\.\d{2})', line)
                if m: meta["closing"] = float(m.group(1).replace(',',''))
                m = re.search(r'MUTASI\s+CR\s*:\s*([\d,]+\.\d{2})', line)
                if m: meta["totalKredit"] = float(m.group(1).replace(',',''))
                m = re.search(r'MUTASI\s+DB\s*:\s*([\d,]+\.\d{2})', line)
                if m: meta["totalDebet"] = float(m.group(1).replace(',',''))

            # ── Parse transaksi ──
            current_tx = None

            for y in sorted(rows.keys()):
                rw    = sorted(rows[y], key=lambda w: w['x0'])
                texts = [w['text'] for w in rw]
                xs    = [w['x0'] for w in rw]
                if not texts: continue

                first_x = xs[0]
                first_t = texts[0]

                # Baris utama transaksi: x<60, diawali DD/MM
                if first_x < 60 and DATE_RE_BCA.match(first_t):
                    if current_tx:
                        transactions.append(current_tx)

                    dd_mm    = first_t
                    rest_str = ' '.join(texts[1:])

                    if 'SALDO' in rest_str and 'AWAL' in rest_str:
                        current_tx = None
                        continue

                    # DB/CR: cari di kolom x 100-470 (bukan x<100 = nama TX)
                    # Kolom DB/CR resmi di x≈149 (TRSF) atau x≈442 (TARIKAN, BI-FAST)
                    # Jangan pakai texts[-1] karena bisa berupa angka saldo
                    dbcr_words = [w['text'] for w in rw
                                  if w['text'] in ('DB', 'CR') and w['x0'] > 100]
                    is_db = 'DB' in dbcr_words

                    # Angka: nominal = angka pertama di x > 350, balance = angka terbesar
                    amt_nums = [w for w in rw if NUM_RE_BCA.match(w['text']) and w['x0'] > 350]
                    if amt_nums:
                        amt_nums_sorted = sorted(amt_nums, key=lambda w: w['x0'])
                        nominal = float(amt_nums_sorted[0]['text'].replace(',',''))
                        balance_words = [w for w in amt_nums_sorted if w['x0'] > 480]
                        balance = float(balance_words[0]['text'].replace(',','')) if balance_words else 0
                    else:
                        nominal = 0
                        balance = 0

                    # Keterangan: teks di kolom keterangan (x < 220)
                    # Buang: angka, DB/CR, tanggal
                    COL_CBG = 220
                    ket_words = [w['text'] for w in rw
                                 if w['x0'] < COL_CBG
                                 and not NUM_RE_BCA.match(w['text'])
                                 and w['text'] not in ('DB', 'CR')
                                 and not DATE_RE_BCA.match(w['text'])]
                    keterangan = ' '.join(ket_words).strip()

                    dd, mm = dd_mm.split('/')
                    yr2    = year[2:] if year else '25'

                    current_tx = {
                        'date'   : f"{dd}/{mm}/{yr2}",
                        'desc'   : keterangan,
                        'debet'  : nominal if is_db else 0,
                        'kredit' : nominal if not is_db else 0,
                        'balance': balance,
                        '_cont'  : [],
                    }

                # Baris lanjutan: x≈170-230
                elif current_tx and 170 < first_x < 240:
                    # Ambil SEMUA teks di baris ini (termasuk angka referensi)
                    line_text = ' '.join(texts).strip()
                    if line_text and line_text != 'Bersambung ke halaman berikut':
                        current_tx['_cont'].append(line_text)

                elif 'Bersambung' in ' '.join(texts):
                    pass

            if current_tx:
                transactions.append(current_tx)
                current_tx = None

    # ── Post-process: gabungkan keterangan + semua baris lanjutan ──
    for tx in transactions:
        cont = tx.pop('_cont', [])

        # Remark lengkap = keterangan utama + semua baris lanjutan
        # Buang baris yang hanya angka duplikat (nomor float tanpa koma ribuan)
        cont_clean = []
        for c in cont:
            # Skip angka duplikat murni (misalnya "135142405.00")
            if re.match(r'^\d+\.\d{2}$', c.strip()):
                continue
            # Skip "Bersambung ke halaman berikut"
            if 'Bersambung' in c:
                continue
            cont_clean.append(c)

        # Gabungkan semua jadi remark lengkap
        parts = [tx['desc']] + cont_clean
        tx['desc'] = ' '.join(p for p in parts if p).strip()

    # Update period
    if transactions and year:
        dates = [t['date'] for t in transactions if t['date']]
        if dates:
            meta['period'] = f"{dates[0]} - {dates[-1]}"

    # Tambah bulan & kategori
    is_idr = meta.get('currency', 'IDR') == 'IDR'
    for tx in transactions:
        tx['month'] = _month_key(tx['date'])
        if is_idr:
            tx['kategori'] = _categorize(tx['desc'], meta['companyName'], tx['debet'], tx['kredit'])
        else:
            tx['kategori'] = 'Non penjualan'

    return meta, transactions



def _is_mandiri_pdf(pdf_path):
    """Deteksi format rekening koran Bank Mandiri."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            words = pdf.pages[0].extract_words()
            text  = ' '.join(w['text'] for w in words[:80])
            return ('Laporan Rekening Koran' in text or 'Account Statement Report' in text) \
                   and ('MCM' in text or 'InhouseTrf' in text)
    except:
        return False


def _parse_pdf_mandiri(pdf_path):
    """
    Parser Bank Mandiri. extract_text(layout=True) posisi kolom:
      col  0-14 = tanggal/detik
      col 15-99 = remark  
      col 100+  = angka debet/kredit/saldo

    Penanda akhir remark: 9910x, angka cabang 14xxx/15xxx, ref panjang >=10 digit.
    """
    meta = {
        "accountNo": "", "companyName": "", "period": "",
        "opening": 0, "totalDebet": 0, "totalKredit": 0, "closing": 0,
        "currency": "IDR"
    }
    transactions = []

    DATE_RE = re.compile(r'^\d{2}/\d{2}/\d{4}')
    AMNT_RE = re.compile(r'[-–]?\s*([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})')
    # Penanda akhir remark: 9910x, angka cabang 14xxx/15xxx (bisa embedded), ref panjang >=10 digit
    END_RE  = re.compile(r'9910[0-9]|1[456]\d{3}|\d{10,}')

    BLN = {"Jan":"01","Feb":"02","Mar":"03","Apr":"04","May":"05","Mei":"05",
           "Jun":"06","Jul":"07","Aug":"08","Agu":"08","Sep":"09","Oct":"10",
           "Okt":"10","Nov":"11","Dec":"12","Des":"12"}

    def _remark(line):
        """Ambil teks kolom remark (pos 15-99), buang angka AMNT."""
        if len(line) <= 15: return ""
        s = line[15:100]
        s = AMNT_RE.sub("", s)
        s = re.sub(r'\s*[-–]\s*$', '', s).strip()
        return s

    def _is_date(line):
        lsp = line.lstrip()
        return bool(DATE_RE.match(lsp)) and (len(line) - len(lsp)) < 10

    def _get_date(line):
        m = DATE_RE.match(line.lstrip())
        if not m: return ""
        d, mo, y = m.group().split('/')
        return f"{d}/{mo}/{y[2:]}"

    def _parse_amnt(line):
        m = AMNT_RE.search(line)
        if not m: return 0.0, 0.0, 0.0
        v = [float(x.replace(',','')) for x in m.groups()]
        if v[0] == 0.0: return 0.0, v[1], v[2]
        if v[1] == 0.0: return v[0], 0.0, v[2]
        return v[0], v[1], v[2]

    # ── Kumpulkan baris dari semua halaman ───────────────────────────────────
    rows = []  # {'type': DATE|AMNT|REMARK, 'date': str, 'remark': str, 'debet':f, 'kredit':f, 'balance':f, 'is_end': bool}

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)

        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text(layout=True, x_density=7, y_density=13)
            if not text: continue
            raw = text.split('\n')

            # Meta
            if page_num == 0:
                full = ' '.join(l.strip() for l in raw if l.strip())
                m = re.search(r'\b(\d{13})\b', full)
                if m: meta['accountNo'] = m.group(1)
                m = re.search(r'\d{13}\s+IDR\s+(.+?)(?:\s{2,}|$)', full)
                if m:
                    nm = m.group(1).strip()
                    meta['companyName'] = nm[:len(nm)//2].strip()
                m = re.search(r'Period\s+(\d{2})\s+(\w+)\s+(\d{4})\s+-\s+(\d{2})\s+(\w+)\s+(\d{4})', full)
                if m:
                    d1,b1,y1,d2,b2,y2 = m.groups()
                    meta['period'] = (f"{d1.zfill(2)}/{BLN.get(b1[:3],'01')}/{y1[2:]} - "
                                      f"{d2.zfill(2)}/{BLN.get(b2[:3],'01')}/{y2[2:]}")
                m = re.search(r'Opening Balance\s*([\d,]+\.\d{2})', full)
                if m: meta['opening'] = float(m.group(1).replace(',',''))
                m = re.search(r'Currency\s+(IDR|USD|EUR|SGD)', full)
                if m: meta['currency'] = m.group(1)

            if page_num == total_pages - 1:
                full = ' '.join(l.strip() for l in raw if l.strip())
                m = re.search(r'Total Amount Debited\s*([\d,]+\.\d{2})', full)
                if m: meta['totalDebet'] = float(m.group(1).replace(',',''))
                m = re.search(r'Total Amount Credited\s*([\d,]+\.\d{2})', full)
                if m: meta['totalKredit'] = float(m.group(1).replace(',',''))
                m = re.search(r'Closing Balance\s*([\d,]+\.\d{2})', full)
                if m: meta['closing'] = float(m.group(1).replace(',',''))

            # Skip header
            tbl = 0
            for i, line in enumerate(raw):
                if 'Posting Date' in line and 'Remark' in line:
                    tbl = i + 1
                    if tbl < len(raw) and raw[tbl].strip() in ('No', 'No.'):
                        tbl += 1
                    break

            for line in raw[tbl:]:
                if not line.strip(): continue
                is_d = _is_date(line)
                is_a = bool(AMNT_RE.search(line))
                rmk  = _remark(line)
                # Untuk baris DATE: remark inline setelah HH:MM: (di posisi 15+)
                # rmk sudah mengambil posisi 15-99 termasuk inline remark setelah jam
                # Tapi jam itu sendiri ada di posisi 15+ juga — perlu dibuang
                if is_d:
                    rmk = re.sub(r'^\d{1,2}:\d{2}:\s*', '', rmk).strip()

                d, k, b = (0.0, 0.0, 0.0)
                if is_a:
                    d, k, b = _parse_amnt(line)

                is_end = bool(END_RE.search(rmk))
                date_s = _get_date(line) if is_d else ""

                typ = 'DATE' if is_d else ('AMNT' if is_a else 'REMARK')
                rows.append({'type': typ, 'date': date_s, 'remark': rmk,
                             'debet': d, 'kredit': k, 'balance': b, 'is_end': is_end})

    # ── Bentuk transaksi ─────────────────────────────────────────────────────
    # Cari semua pasangan DATE → AMNT
    pairs = []  # (date_idx, amnt_idx, date_str)
    i = 0
    while i < len(rows):
        if rows[i]['type'] == 'DATE':
            d_idx = i
            date_s = rows[i]['date']
            # Cari AMNT berikutnya
            for j in range(i+1, min(i+5, len(rows))):
                if rows[j]['type'] == 'AMNT':
                    pairs.append((d_idx, j, date_s))
                    i = j + 1
                    break
            else:
                i += 1
        else:
            i += 1

    # Untuk setiap TX:
    # - Pre-remark: baris REMARK setelah END_MARKER TX sebelumnya sampai DATE TX ini
    #   (Hanya jika TX sebelumnya punya END_MARKER)
    # - Inline DATE: remark di baris DATE TX ini
    # - Inline AMNT: remark di baris AMNT TX ini
    # - Post-remark: baris REMARK setelah AMNT sampai END_MARKER (inklusif)

    for idx, (d_idx, a_idx, date_s) in enumerate(pairs):
        d, k, b = rows[a_idx]['debet'], rows[a_idx]['kredit'], rows[a_idx]['balance']
        if d == 0 and k == 0: continue

        # Cari END_MARKER TX sebelumnya
        prev_end_idx = -1
        if idx > 0:
            prev_a = pairs[idx-1][1]
            for p in range(prev_a, d_idx):
                if rows[p]['is_end']:
                    prev_end_idx = p  # ambil yang terakhir

        # Pre-remark
        pre = []
        if idx == 0:
            # TX pertama: ambil semua REMARK sebelum DATE pertama
            for p in range(0, d_idx):
                if rows[p]['type'] == 'REMARK' and rows[p]['remark']:
                    pre.append(rows[p]['remark'])
        elif prev_end_idx >= 0:
            # Ada END_MARKER sebelumnya: pre dimulai setelah END_MARKER
            for p in range(prev_end_idx + 1, d_idx):
                if rows[p]['type'] == 'REMARK' and rows[p]['remark']:
                    pre.append(rows[p]['remark'])
        else:
            # Tidak ada END_MARKER: scan mundur dari d_idx untuk REMARK langsung sebelum DATE
            # (biasanya 1-2 baris yang merupakan pre-remark TX ini)
            tmp = []
            for p in range(d_idx - 1, max(-1, pairs[idx-1][1]-1) if idx > 0 else -1, -1):
                if rows[p]['type'] == 'REMARK' and rows[p]['remark']:
                    tmp.insert(0, rows[p]['remark'])
                elif rows[p]['type'] in ('DATE', 'AMNT'):
                    break
            pre = tmp

        # Inline DATE dan between
        date_inline = rows[d_idx]['remark'] if rows[d_idx]['remark'] else ""
        between = [rows[p]['remark'] for p in range(d_idx+1, a_idx)
                   if rows[p]['type'] == 'REMARK' and rows[p]['remark']]

        # Inline AMNT
        amnt_inline = rows[a_idx]['remark'] if rows[a_idx]['remark'] else ""

        # Post-remark: setelah AMNT sampai END_MARKER TERAKHIR sebelum DATE berikutnya
        # (bukan berhenti di END_MARKER pertama, karena bisa ada beberapa baris setelahnya)
        next_d = pairs[idx+1][0] if idx+1 < len(pairs) else len(rows)
        post_all = []
        last_end_in_post = -1
        for p in range(a_idx + 1, next_d):
            if rows[p]['type'] == 'DATE': break
            if rows[p]['remark']:
                post_all.append((p, rows[p]['remark']))
            if rows[p]['is_end']:
                last_end_in_post = len(post_all) - 1  # index di post_all

        # Ambil hanya sampai END_MARKER terakhir (inklusif)
        if last_end_in_post >= 0:
            post = [r for _, r in post_all[:last_end_in_post + 1]]
        else:
            # Tidak ada END_MARKER di post → ambil semua (TX tanpa penanda akhir)
            post = [r for _, r in post_all]

        parts = pre + ([date_inline] if date_inline else []) + between + \
                ([amnt_inline] if amnt_inline else []) + post
        remark = ' '.join(p for p in parts if p).strip()
        remark = re.sub(r'  +', ' ', remark)

        transactions.append({
            'date'   : date_s,
            'desc'   : remark,
            'debet'  : d,
            'kredit' : k,
            'balance': b,
        })

    # Tambah bulan & kategori
    is_idr = meta.get('currency', 'IDR') == 'IDR'
    for tx in transactions:
        tx['month'] = _month_key(tx['date'])
        if is_idr:
            tx['kategori'] = _categorize(tx['desc'], meta['companyName'], tx['debet'], tx['kredit'])
        else:
            tx['kategori'] = 'Non penjualan'

    return meta, transactions


def _is_bni_pdf(pdf_path):
    """Deteksi format rekening koran Bank BNI."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            words = pdf.pages[0].extract_words()
            text = ' '.join(w['text'] for w in words[:60])
            return 'ACCOUNT STATEMENT' in text and 'DB/CR' in text and \
                   ('Account No.' in text or 'Account Type' in text)
    except:
        return False


def _parse_pdf_bni(pdf_path):
    """
    Parser rekening koran Bank BNI Account Statement.

    Kolom (x position):
      x < 50    = Posting Date (DD/MM/YYYY)
      x 50-230  = Effective Date
      x 230-300 = Branch (multi-baris)
      x ~298    = Journal number
      x > 330   = Transaction Description (multi-baris)
      x 480-570 = Amount (bold PDF duplikat karakter)
      x 570-610 = DB/CR (K/D)
      x > 610   = Balance

    Amount: PDF bold rendering menyebabkan karakter overlap di posisi x sama.
    Solusi: group chars per x, untuk overlap ambil char non-'0'.

    Struktur per TX:
      [desc pre-TX]    x>330, sebelum baris tanggal
      TX baris         x<50, DD/MM/YYYY, berisi Journal, DB/CR, Balance
      [desc post-TX]   x>330, setelah baris tanggal
    """
    meta = {
        "accountNo": "", "companyName": "", "period": "",
        "opening": 0, "totalDebet": 0, "totalKredit": 0, "closing": 0,
        "currency": "IDR"
    }
    transactions = []

    DATE_RE = re.compile(r'^\d{2}/\d{2}/\d{4}$')
    NUM_RE  = re.compile(r'^[\d,]+\.\d{2}$')

    BLN = {"Jan":"01","Feb":"02","Mar":"03","Apr":"04","May":"05","Mei":"05",
           "Jun":"06","Jul":"07","Aug":"08","Agu":"08","Sep":"09","Oct":"10",
           "Okt":"10","Nov":"11","Dec":"12","Des":"12"}

    X_DESC = 330   # description mulai di sini
    X_AMT  = 479   # amount mulai di sini
    X_DBCR = 570   # DB/CR
    X_BAL  = 610   # balance

    def _bni_date(text):
        try:
            d, mo, y = text.split('/')
            return f"{d}/{mo}/{y[2:]}"
        except:
            return ""

    def _decode_amount(chars_in_range):
        """
        Decode amount dari chars dengan x positions.
        Bold PDF: dua char di posisi x sama → ambil non-'0'.
        """
        by_x = {}
        for c in chars_in_range:
            xk = round(c['x0'] * 10) / 10
            by_x.setdefault(xk, []).append(c['text'])

        result = ''
        for xk in sorted(by_x.keys()):
            clist = by_x[xk]
            if len(clist) == 1:
                result += clist[0]
            else:
                non_zero = [c for c in clist if c != '0']
                result += non_zero[0] if non_zero else clist[0]

        cleaned = re.sub(r'\.\.+', '.', result)
        cleaned = re.sub(r',,+',   ',', cleaned)
        if NUM_RE.match(cleaned):
            return float(cleaned.replace(',', ''))
        return 0.0

    def _is_valid_bal(text):
        """Angka valid untuk balance (bukan jam HH.MM.SS)."""
        return bool(NUM_RE.match(text.strip()))

    SKIP = {'Posting','Effective','Branch','Journal','Transaction','Description',
            'Amount','DB/CR','Balance','Ending','Total','Ledger','Account',
            'Period','Page','ACCOUNT','Information','STATEMENT','Type','CURRENT',
            'No.','PT(IDR)','Debet','Credit'}

    # ── Kumpulkan semua data ─────────────────────────────────────────────────
    all_rows  = []   # list of (gy, date_str, dbcr, balance, desc_words)
    amt_by_gy = {}   # gy → amount dari chars

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)

        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            chars = page.chars

            # ── Meta ────────────────────────────────────────────────────────
            if page_num == 0:
                # Gabung semua teks halaman 1 per baris, gunakan regex
                rows_p1 = {}
                for w in words:
                    y = round(w['top']/2)*2
                    rows_p1.setdefault(y, []).append(w)
                full_p1 = ' '.join(
                    ' '.join(w['text'] for w in sorted(rows_p1.get(y,[]), key=lambda w: w['x0']))
                    for y in sorted(rows_p1.keys())
                )
                if not meta['accountNo']:
                    # Format 1: "Account No. : 0045206873"
                    m = re.search(r'Account\s+No\.?\s*:\s*(\d+)', full_p1)
                    # Format 2: "45206873 / PANGAN LESTARI PT(IDR)"
                    if not m:
                        m = re.search(r'(\d{6,12})\s*/\s*[A-Z]', full_p1)
                    if m: meta['accountNo'] = m.group(1)
                if not meta['companyName']:
                    m = re.search(r'(\d{6,12})\s*/\s*(.+?)(?:\s*\(IDR\)|\s{2,}|$)', full_p1)
                    if m: meta['companyName'] = m.group(2).strip()
                if not meta['period']:
                    m = re.search(r'Period\s*:\s*(\d{2}-\w{3}-\d{2})\s*-\s*(\d{2}-\w{3}-\d{2})', full_p1)
                    if m:
                        def bd(s):
                            p = s.strip().split('-')
                            return f"{p[0].zfill(2)}/{BLN.get(p[1][:3],'01')}/{p[2]}" if len(p)==3 else s
                        meta['period'] = f"{bd(m.group(1))} - {bd(m.group(2))}"
                # Ledger Balance (saldo awal): cari angka setelah "Ledger Balance"
                if not meta['opening']:
                    m = re.search(r'Ledger\s+Balance[:\s]+([\d,]+\.\d{2})', full_p1)
                    if m: meta['opening'] = float(m.group(1).replace(',',''))

            # ── Summary ─────────────────────────────────────────────────────
            if page_num == total_pages - 1:
                rows_last = {}
                for w in words:
                    y = round(w['top']/2)*2
                    rows_last.setdefault(y, []).append(w)
                full_last = ' '.join(
                    ' '.join(w['text'] for w in sorted(rows_last.get(y,[]), key=lambda w: w['x0']))
                    for y in sorted(rows_last.keys())
                )
                # Ending Balance (saldo akhir)
                m = re.search(r'Ending\s+Balance\s*:\s*([\d,]+\.\d{2})', full_last)
                if m: meta['closing'] = float(m.group(1).replace(',',''))
                # Total Debet
                m = re.search(r'Total\s+Debet\s*:\s*\d*\s*([\d,]+\.\d{2})', full_last)
                if m: meta['totalDebet'] = float(m.group(1).replace(',',''))
                # Total Credit
                m = re.search(r'Total\s+Credit\s*:\s*\d*\s*([\d,]+\.\d{2})', full_last)
                if m: meta['totalKredit'] = float(m.group(1).replace(',',''))

            # ── Amount per baris (chars-based, exact) ────────────────────────
            chars_by_y = {}
            for c in chars:
                if X_AMT < c['x0'] < X_DBCR:
                    y = round(c['top']/2)*2
                    chars_by_y.setdefault(y, []).append(c)

            for y, clist in chars_by_y.items():
                v = _decode_amount(clist)
                if v > 0:
                    gy = page_num * 100000 + y
                    amt_by_gy[gy] = v

            # ── Baris tabel (words) ──────────────────────────────────────────
            page_rows = {}
            for w in words:
                y = round(w['top']/2)*2
                page_rows.setdefault(y, []).append(w)

            header_y = 0
            for y in sorted(page_rows.keys()):
                texts = [w['text'] for w in page_rows[y]]
                if 'Posting' in texts and 'DB/CR' in texts:
                    header_y = y
                    break

            for y in sorted(page_rows.keys()):
                if y <= header_y: continue
                rw = sorted(page_rows[y], key=lambda w: w['x0'])
                texts = [w['text'] for w in rw]
                xs    = [w['x0']   for w in rw]
                if not texts: continue
                if any(t in SKIP for t in texts[:3]): continue

                gy = page_num * 100000 + y

                # TX baris: DD/MM/YYYY di x < 50
                is_tx = xs[0] < 50 and DATE_RE.match(texts[0])
                date_str = _bni_date(texts[0]) if is_tx else ""

                # DB/CR
                dbcr = ""
                for w in rw:
                    if X_DBCR <= w['x0'] < X_BAL and w['text'] in ('K', 'D'):
                        dbcr = w['text']
                        break

                # Balance
                bal = [w['text'] for w in rw if w['x0'] >= X_BAL and _is_valid_bal(w['text'])]
                balance = float(bal[-1].replace(',','')) if bal else 0.0

                # Desc words: x > 330, x < 570, bukan angka/K/D
                desc_words = []
                for w in rw:
                    if w['x0'] <= X_DESC: continue
                    if w['x0'] >= X_DBCR: continue
                    t = w['text']
                    if t in ('K', 'D'): continue
                    # Skip angka yang merupakan amount (x > X_AMT)
                    if w['x0'] >= X_AMT and _is_valid_bal(t): continue
                    # Skip angka bold-duplicate
                    if w['x0'] >= X_AMT and re.search(r'\.\.|,,', t): continue
                    desc_words.append(t)

                all_rows.append({
                    'type'   : 'TX' if is_tx else 'RMK',
                    'gy'     : gy,
                    'date'   : date_str,
                    'desc'   : desc_words,
                    'dbcr'   : dbcr,
                    'balance': balance,
                })

    # ── Bentuk transaksi ─────────────────────────────────────────────────────
    tx_idx = [i for i, r in enumerate(all_rows) if r['type'] == 'TX']

    for k, ti in enumerate(tx_idx):
        row = all_rows[ti]
        if not row['date']: continue

        dbcr    = row['dbcr']

        prev_ti = tx_idx[k-1] if k > 0 else -1
        next_ti = tx_idx[k+1] if k+1 < len(tx_idx) else len(all_rows)

        pre_rows  = [all_rows[j] for j in range(prev_ti+1, ti)  if all_rows[j]['type'] == 'RMK']
        post_rows = [all_rows[j] for j in range(ti+1, next_ti)  if all_rows[j]['type'] == 'RMK']

        # Balance: cari di seluruh segmen (pre + TX baris + post)
        # Balance selalu di x >= X_BAL (610), bisa di baris pre-desc atau baris TX
        balance = row['balance']  # coba dari baris TX dulu
        if balance == 0:
            # Cari di pre_rows (bisa ada di baris deskripsi terakhir sebelum TX)
            for r in reversed(pre_rows):
                if r['balance'] > 0:
                    balance = r['balance']
                    break
        if balance == 0:
            # Cari di post_rows
            for r in post_rows:
                if r['balance'] > 0:
                    balance = r['balance']
                    break

        # Amount: cari dari chars-based amt_by_gy
        # Pre-baris biasanya punya amount (untuk TRANSFER DARI/KE panjang)
        amount = 0.0
        for r in reversed(pre_rows):
            v = amt_by_gy.get(r['gy'], 0)
            if v > 0:
                amount = v
                break
        # Fallback: TX baris sendiri
        if amount == 0:
            v = amt_by_gy.get(row['gy'], 0)
            if v > 0:
                amount = v
        # Fallback: post baris
        if amount == 0:
            for r in post_rows:
                v = amt_by_gy.get(r['gy'], 0)
                if v > 0:
                    amount = v
                    break

        if amount == 0: continue

        debet  = amount if dbcr == 'D' else 0.0
        kredit = amount if dbcr == 'K' else 0.0
        if debet == 0 and kredit == 0: continue

        # Deskripsi TX BNI:
        # Pre-rows mengandung campuran post-desc TX sebelumnya + pre-desc TX ini.
        # Batas: pre-desc TX ini dimulai dari kata kunci BNI (TRANSFER, SETOR, TARIK, dll).
        # Scan pre_rows dari belakang: ambil sampai bertemu kata kunci di baris sebelumnya.
        BNI_ANCHORS = {'TRANSFER', 'SETOR', 'TARIK', 'JASA', 'BIAYA', 'PPH',
                       'INKASO', 'KOREKSI', 'AUTO', 'SETORAN', 'PENARIKAN'}

        # Cari titik "awal" pre-desc TX ini di pre_rows
        # Baris yang mengandung anchor kata kunci = awal TX ini
        anchor_idx = len(pre_rows)  # default: tidak ambil pre_rows
        for pi in range(len(pre_rows)):
            words_pi = pre_rows[pi]['desc']
            if any(w.upper() in BNI_ANCHORS for w in words_pi):
                anchor_idx = pi
                break

        # Jika TX ini sudah punya desc inline (JASA GIRO, PPH, BIAYA, SETOR, TARIK),
        # skip pre_rows sepenuhnya (sudah cukup dari inline)
        tx_has_inline = bool(row['desc'])
        if tx_has_inline and anchor_idx == len(pre_rows):
            # Tidak ada anchor di pre → pre milik TX sebelumnya, skip semua
            pre_desc = []
        else:
            pre_desc  = pre_rows[anchor_idx:]

        # Post-rows: stop saat menemukan anchor TX berikutnya
        # (baris yang mengandung kata kunci BNI deskripsi = awal TX berikutnya)
        post_desc = []
        for r in post_rows:
            words_r = r['desc']
            # Jika baris ini mengandung anchor dan bukan baris pertama post → stop
            if post_desc and any(w.upper() in BNI_ANCHORS for w in words_r):
                break
            post_desc.append(r)

        all_desc = []
        for r in pre_desc:
            all_desc.extend(r['desc'])
        all_desc.extend(row['desc'])
        for r in post_desc:
            all_desc.extend(r['desc'])

        desc = ' '.join(w for w in all_desc if w).strip()
        desc = re.sub(r'\s+', ' ', desc)

        transactions.append({
            'date'   : row['date'],
            'desc'   : desc,
            'debet'  : debet,
            'kredit' : kredit,
            'balance': balance,
        })

    # ── Perbaiki balance yang kosong (0) ────────────────────────────────────
    # Strategi:
    # 1. Jika balance TX ada → pakai
    # 2. Jika tidak, hitung manual: running balance dari opening
    # 3. TX terakhir: gunakan Ending Balance dari meta jika tersedia
    if transactions:
        # Pass 1: isi balance yang masih 0 dengan perhitungan running balance
        running = meta.get('opening', 0)
        for tx in transactions:
            if tx['balance'] > 0:
                running = tx['balance']
            else:
                # Hitung manual
                running = running + tx['kredit'] - tx['debet']
                tx['balance'] = round(running, 2)

        # Pass 2: TX terakhir gunakan Ending Balance jika lebih akurat
        if meta.get('closing', 0) > 0:
            transactions[-1]['balance'] = meta['closing']

    # Tambah bulan & kategori
    is_idr = meta.get('currency', 'IDR') == 'IDR'
    for tx in transactions:
        tx['month'] = _month_key(tx['date'])
        if is_idr:
            tx['kategori'] = _categorize(tx['desc'], meta['companyName'], tx['debet'], tx['kredit'])
        else:
            tx['kategori'] = 'Non penjualan'

    return meta, transactions


def parse_pdf(pdf_path):
    """Auto-detect format BRI, BCA, Mandiri, atau BNI, lalu parse."""
    if _is_bca_pdf(pdf_path):
        return _parse_pdf_bca(pdf_path)
    if _is_mandiri_pdf(pdf_path):
        return _parse_pdf_mandiri(pdf_path)
    if _is_bni_pdf(pdf_path):
        return _parse_pdf_bni(pdf_path)
    return _parse_pdf_bri(pdf_path)


def _parse_pdf_bri(pdf_path):

    meta = {
        "accountNo": "", "companyName": "", "period": "",
        "opening": 0, "totalDebet": 0, "totalKredit": 0, "closing": 0,
        "currency": "IDR"
    }
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        n_pages = len(pdf.pages)

        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            rows  = defaultdict(list)
            for w in words:
                rows[round(w['top'] / 2) * 2].append(w)

            # ── Deteksi kolom dari header tabel ──
            cols = detect_columns(rows)

            # ── Temukan Y baris summary ──
            summary_y = find_summary_y(rows)

            # ── Meta dari halaman 1 ──
            if page_num == 0:
                for y in sorted(rows.keys()):
                    rw   = sorted(rows[y], key=lambda w: w['x0'])
                    line = ' '.join(w['text'] for w in rw)

                    # No. Rekening: ambil angka 13-16 digit setelah ':' atau 'Rekening'
                    if not meta["accountNo"]:
                        m = re.search(r'(?:Rekening|Nomor|No\.?)[\s:]+([\d]{13,16})\b', line, re.I)
                        if not m:
                            m = re.search(r':\s*(\d{13,16})\b', line)
                        if m: meta["accountNo"] = m.group(1)

                    # Nama perusahaan: baris yang mengandung nama tapi bukan header
                    # Ambil teks di sisi kiri (x < 300) yang bukan keyword
                    if not meta["companyName"] and y > 100 and y < 200:
                        left_words = [w['text'] for w in rw
                                      if w['x0'] < 300
                                      and w['text'] not in ('Kepada','Yth.','/','To',':')
                                      and not re.match(r'^\d', w['text'])]
                        candidate = ' '.join(left_words).strip()
                        # Harus minimal 3 karakter dan semua huruf besar (nama perusahaan)
                        if (len(candidate) >= 3
                                and candidate.replace(' ','').isupper()
                                and candidate not in ('RAYA','JL','LT','RT')):
                            meta["companyName"] = candidate

                    # Periode transaksi
                    pm = re.search(r'(\d{2}/\d{2}/\d{2})\s*-\s*(\d{2}/\d{2}/\d{2})', line)
                    if pm and not meta["period"]:
                        meta["period"] = f"{pm.group(1)} - {pm.group(2)}"

                    # Deteksi rekening pinjaman: ada kata "PINJAMAN" atau "Plafond"
                    if not meta.get("is_pinjaman") and re.search(r'PINJAMAN|Plafond', line):
                        meta["is_pinjaman"] = True

                    # Valuta — hanya dari baris header "Valuta : USD" atau "Currency : USD"
                    if meta["currency"] == "IDR" and re.search(r'\b(Valuta|Currency)\b', line, re.I):
                        cm = re.search(r'\b(USD|EUR|SGD|CNY|JPY|GBP|AUD)\b', line)
                        if cm:
                            meta["currency"] = cm.group(1)

            # ── Summary (Saldo Awal dst.) ──
            for y in sorted(rows.keys()):
                rw   = sorted(rows[y], key=lambda w: w['x0'])
                nums = [w for w in rw if NUM_RE.match(w['text'])]
                if len(nums) >= 4:
                    vs = [float(n['text'].replace(',', '')) for n in nums]
                    if vs[1] > 1e5:  # Total Debet pasti besar
                        meta["opening"]     = vs[0]
                        meta["totalDebet"]  = vs[1]
                        meta["totalKredit"] = vs[2]
                        meta["closing"]     = vs[3]

            # ── Parse transaksi ──
            current_tx = None

            for y in sorted(rows.keys()):
                # STOP: jangan proses di atas / sama dengan baris summary
                if y >= summary_y:
                    break

                row_words = sorted(rows[y], key=lambda w: w['x0'])
                full_text = ' '.join(w['text'] for w in row_words)

                # Skip baris header/info
                if any(s in full_text for s in SKIP_KW):
                    continue
                # Skip baris footer/stop
                if any(s in full_text for s in STOP_KW):
                    break

                first  = row_words[0]['text'] if row_words else ''
                is_tx  = bool(DATE_RE.match(first))

                if is_tx:
                    if current_tx:
                        transactions.append(current_tx)
                    date_str = ''
                    desc_parts = []
                    debet = kredit = balance = 0.0

                    for w in row_words:
                        t = w['text']; x = w['x0']
                        if DATE_RE.match(t) and not date_str:
                            date_str = t
                        elif TIME_RE.match(t) and date_str and ' ' not in date_str:
                            date_str += ' ' + t
                        elif TELLER_RE.match(t):
                            pass  # skip teller ID
                        elif NUM_RE.match(t):
                            v = float(t.replace(',', ''))
                            if   x >= cols['saldo']:   balance = v
                            elif x >= cols['kredit']:  kredit  = v
                            elif x >= cols['debet']:   debet   = v
                            # angka di kiri cols['debet'] = bagian deskripsi (no. rekening, dll)
                        elif x < cols['teller']:
                            # Bagian deskripsi
                            if not DATE_RE.match(t) and not TIME_RE.match(t):
                                desc_parts.append(t)

                    current_tx = {
                        'date': date_str,
                        'desc': ' '.join(desc_parts),
                        'debet': debet, 'kredit': kredit, 'balance': balance
                    }

                elif current_tx:
                    # Baris lanjutan deskripsi — hanya jika tidak ada angka di kolom kanan
                    right_nums = [w for w in row_words
                                  if NUM_RE.match(w['text']) and w['x0'] >= cols['debet']]
                    if not right_nums:
                        extra = ' '.join(
                            w['text'] for w in row_words
                            if w['x0'] < cols['teller']
                            and not TELLER_RE.match(w['text'])
                            and not NUM_RE.match(w['text'])
                        ).strip()
                        if extra:
                            current_tx['desc'] = (current_tx['desc'] + ' ' + extra).strip()

            if current_tx:
                transactions.append(current_tx)
                current_tx = None

    # Tambah bulan & kategori
    is_idr  = meta.get('currency', 'IDR') == 'IDR'
    is_pinj = meta.get('is_pinjaman', False)
    for j, tx in enumerate(transactions):
        tx['month'] = _month_key(tx['date'])

        if is_pinj and is_idr:
            up  = (tx['desc'] or '').upper()
            prev_tx = transactions[j-1] if j > 0 else None
            prev_kr = prev_tx['kredit'] if prev_tx else 0

            # Aturan 1: RTGS dari nama pihak lain = Penjualan
            is_rtgs_lain = (
                up.startswith('RTGS#')
                and tx['kredit'] > 0
                and not _contains_own_name(tx['desc'], meta['companyName'])
                and not re.match(r'^RTGS#[^A-Z]', up)
                and len(up.replace('RTGS#','').strip()) > 3
            )

            # Aturan 2: *BAYAR POKOK dengan nominal BERBEDA dari baris sebelumnya
            # = LLG dari customer (tidak ada transaksi pasangan di atasnya)
            is_bayar_llg = (
                '*BAYAR POKOK' in up
                and tx['kredit'] > 0
                and abs(tx['kredit'] - prev_kr) > 1  # nominal berbeda
            )

            if is_rtgs_lain or is_bayar_llg:
                tx['kategori'] = 'Penjualan'
            else:
                tx['kategori'] = 'Non penjualan'

        elif not is_idr:
            tx['kategori'] = 'Non penjualan'
        else:
            tx['kategori'] = _categorize(tx['desc'], meta['companyName'], tx['debet'], tx['kredit'])
    return meta, transactions

def _month_key(d):
    m = re.match(r'^(\d{2})/(\d{2})/(\d{2})', d or '')
    if not m: return 'Unknown'
    return f"{MONTHS_ID[int(m.group(2))-1]} 20{m.group(3)}"


def _contains_own_name(desc, company_name):
    up_desc = (desc or '').upper()
    up_name = company_name.upper()

    # Hapus prefix badan hukum yang tidak spesifik
    PREFIX = {'PT', 'CV', 'UD', 'PD', 'TB', 'TBK', 'TBKK', 'AN', 'THE'}
    all_words = up_name.split()
    name_words = [w for w in all_words if w not in PREFIX]

    # 1. Nama lengkap
    if up_name in up_desc:
        return True

    # 2. Kombinasi 2 kata berurutan (minimal 1 kata >= 4 huruf, total >= 7 karakter)
    for i in range(len(name_words) - 1):
        w1, w2 = name_words[i], name_words[i+1]
        if max(len(w1), len(w2)) >= 4:
            bigram = w1 + ' ' + w2
            if bigram in up_desc:
                return True

    # 3. Nama terpotong: cek "KATA1 KATA2_AWALAN" (nama dipotong di tengah kata)
    #    Contoh: "NUGA SIGMA POTENZIA" -> cek "NUGA SIGMA P", "NUGA SIGMA PO", dst.
    for i in range(len(name_words) - 1):
        w1, w2 = name_words[i], name_words[i+1]
        if len(w1) >= 4:
            for trunc in range(1, len(w2)):  # SIGMA P, SIGMA PO, SIGMA POT, ...
                pattern = w1 + ' ' + w2[:trunc]
                # Pastikan pattern diikuti spasi/non-huruf (bukan bagian kata lain)
                import re
                if re.search(re.escape(pattern) + r'(?:\s|$|[^A-Z])', up_desc):
                    return True

    # 4. Kata tunggal sangat spesifik (>= 8 huruf, bukan kata generik)
    GENERIC = {
        'MANDIRI','BERSAMA','SEJAHTERA','INDONESIA','UTAMA','JAYA','MAKMUR',
        'ABADI','MAJU','SENTOSA','PRIMA','NUSANTARA','PERSADA','SARANA',
        'PRATAMA','PERDANA','SETIA','KARYA','PUTRA','PUTRI','ANDALAN',
        'SUKSES','MULIA','LANCAR','AMANAH','BERKAH','INDAH','AGUNG',
    }
    for w in name_words:
        if len(w) >= 8 and w not in GENERIC and w in up_desc:
            return True

    return False

def _is_company_like(name_part):
    """
    Apakah suatu string terlihat seperti nama badan usaha / instansi
    (bukan nama orang, bukan nomor rekening, bukan kode sistem)?
    """
    # Tanda badan usaha
    ENTITY_MARKERS = ['PT','CV','UD','PD','BANK','DINAS','KEMENTERIAN',
                      'PEMKOT','PEMKAB','PEMPROV','BRI','BNI','MANDIRI',
                      'KOPERASI','YAYASAN','BPJS','KAI','PLN','PERTAMINA',
                      'TELKOM','BULOG','PERUM','PERSEROAN','TBKK','TBK']
    up = name_part.upper()
    for m in ENTITY_MARKERS:
        if m in up:
            return True
    # Lebih dari 2 kata huruf besar semua → kemungkinan nama perusahaan/instansi
    words = [w for w in up.split() if w.isalpha() and len(w) >= 3]
    if len(words) >= 2 and all(w == w.upper() for w in words):
        return True
    return False

# Nama kota / entitas generik yang muncul setelah RTGS# → bukan customer nyata
RTGS_NON_PENJ_NAMES = set()  # tidak dipakai lagi

def _make_abbreviations(company_name):
    PREFIX = {'PT','CV','UD','PD','TB','TBK','AN','THE'}
    words = [w for w in company_name.upper().split() if w not in PREFIX and len(w) >= 2]
    abbrevs = set()
    if not words: return abbrevs
    abbrevs.add(''.join(w[0] for w in words))
    if len(words) >= 2: abbrevs.add(words[0][0] + words[1][0])
    if len(words) >= 2: abbrevs.add(words[-2][0] + words[-1][0])
    return {a for a in abbrevs if len(a) >= 2}

def _extract_customer_name(desc):
    """
    Ekstrak nama customer dari deskripsi transaksi.
    Format BRI:    NBMB JEMMY BUDYANTO TO PANGAN LESTARI ESB:...   → JEMMY BUDYANTO
    Format BCA:    TRSF E-BANKING ... CAHAYA SETIA UTAMA            → Cahaya Setia Utama
    Format Mandiri: MCM InhouseTrf DARI NURUL KURNIAWATI ...        → Nurul Kurniawati
    Format BNI:    TRANSFER DARI | PEMINDAHAN DARI 251020101 LION SUPERINDO PT | ...
                   SETOR TUNAI | MITRA BUANA NIAGA                  → Mitra Buana Niaga
    """
    import re
    up = desc.strip().upper()

    # ── FORMAT BNI ──────────────────────────────────────────────────────────
    # Pola: "TRANSFER DARI | PEMINDAHAN DARI <kode> <NAMA> | ..."
    # Atau:  "SETOR TUNAI | <NAMA>"
    if up.startswith('TRANSFER DARI') or up.startswith('SETOR TUNAI') or up.startswith('PEMINDAHAN DARI'):
        # SETOR TUNAI: nama setelah pipe
        m = re.search(r'SETOR TUNAI\s*\|\s*(.+?)(?:\s*\||\s*$)', desc, re.I)
        if m:
            name = m.group(1).strip()
            name = re.sub(r'\s+', ' ', name)
            if name and not re.match(r'^[\d\s]+$', name):
                return name.title()

        # TRANSFER DARI: nama setelah kode rekening di "PEMINDAHAN DARI <kode> <NAMA>"
        # Format: "PEMINDAHAN DARI 251020101 LION SUPERINDO PT"
        m = re.search(r'PEMINDAHAN\s+DARI\s+\d+\s+(.+?)(?:\s*\||\s*$)', desc, re.I)
        if m:
            name = m.group(1).strip()
            # Buang kode referensi di akhir (angka panjang)
            name = re.sub(r'\s*\d{6,}.*$', '', name).strip()
            name = re.sub(r'\s+', ' ', name).strip()
            if name and not re.match(r'^[\d\s]+$', name) and len(name) >= 3:
                return name.title()

        return 'Transfer BNI' 

    # ── FORMAT BCA: nama customer di akhir remark ───────────────────────────
    if re.match(r'^(TRSF E-BANKING|KR OTOMATIS|SWITCHING|BYR VIA E-BANKING|SETORAN|BI-FAST)', up):

        OWN_W = {'PANGAN','LESTARI','PANGLES','PANGANLES','AUTOCR-PL',
                 'PANGA','LESTA','LESTAR','PANGANL'}

        # ── SETORAN TUNAI ──────────────────────────────────────────────────
        if up.startswith('SETORAN TUNAI'):
            rest = re.sub(r'^SETORAN TUNAI\s+', '', desc, flags=re.I).strip()
            # Kode BCA/xxx/(KOTA) → ambil kota/nama di dalam kurung jika bukan kota sendiri
            # Jika tidak ada kode → nama langsung
            m = re.match(r'^BCA[/\d\s]+/[\d/\s]+(?:\(.*?\))?\s*([\w\s]+)$', rest, re.I)
            if m:
                name = m.group(1).strip()
                if name and len(name) > 3:
                    return name.title()
            # Bersihkan kode di awal
            rest2 = re.sub(r'^[A-Z]{2,3}[\d/\-\s]+(?:\(.*?\))?\s*', '', rest).strip()
            if rest2 and not re.match(r'^[\d/]', rest2) and len(rest2) > 3:
                # Hapus nama kota BCA di akhir
                rest2 = re.sub(r'\s+(?:JKT|SBY|BANDUNG|JAKARTA|SURABAYA|BCA\s+\w+)\s*$', '', rest2, flags=re.I).strip()
                if rest2 and len(rest2) > 3:
                    return rest2[:30].title()
            return 'Setoran Tunai'

        # ── SWITCHING ─────────────────────────────────────────────────────
        if up.startswith('SWITCHING'):
            rest = re.sub(r'^SWITCHING\s+(?:CR|DR)\s+\w+\s+', '', up).strip()
            rest = re.sub(r'/[\d\.]+.*$', '', rest).strip()
            rest = re.sub(r'^/', '', rest).strip()
            if rest and len(rest) > 3 and not re.match(r'^[/\d\.]', rest):
                return rest[:30].title()
            return None

        # ── KR OTOMATIS LLG / RTGS ────────────────────────────────────────
        if up.startswith('KR OTOMATIS LLG') or up.startswith('KR OTOMATIS RTGS'):
            # Format: "KR OTOMATIS LLG-BANK NAMA [kode referensi akhir]"
            rest = re.sub(r'^KR OTOMATIS (?:LLG|RTGS)-\S+\s*', '', up).strip()
            # Bersihkan kode referensi umum di akhir
            rest = re.sub(r'\s+[-–]\d{4,}.*$', '', rest).strip()     # -0183971881
            rest = re.sub(r'\s+PCM[\dA-Z]+.*$', '', rest).strip()     # PCM0183971881
            rest = re.sub(r'\s+\d{8,}[-\w]*\s*$', '', rest).strip()  # 20250108-ID875
            rest = re.sub(r'\s+\d{6,}\s*$', '', rest).strip()         # 090125
            rest = re.sub(r'\s+\|\s*$', '', rest).strip()
            rest = re.sub(r'\s+[-–]\s*$', '', rest).strip()
            # Bersihkan kode invoice / nomor surat di akhir: SO0124..., MCM-2501...
            rest = re.sub(r'\s+(?:SO|MCM|REF|TT|PLP|PCM|CK|INV|PYMNT)\w*[\d\-]+.*$', '', rest, flags=re.I).strip()
            # Bersihkan teks nama sendiri yang masuk
            words = rest.split()
            clean = [w for w in words 
                     if w not in OWN_W
                     and not re.match(r'^[0-9,]+$', w)
                     and not re.match(r'^PYMNT$', w, re.I)]
            # Ambil max 5 kata
            name = ' '.join(clean[:5]).strip()
            # Bersihkan kode invoice multi-kata yang tersisa
            name = re.sub(r'\s+(?:RI|S)\d{10,}.*$', '', name).strip()
            if name and len(name) >= 3:
                return name.title()
            return None

        # ── TRSF E-BANKING / BYR VIA / BI-FAST BIF ──────────────────────
        # Format: TRSF E-BANKING DDMM/FTSCY/WS95xxx [noise] [NAMA CUSTOMER]
        #         BI-FAST BIF TRANSFER NNN [noise] [NAMA CUSTOMER]
        # Nama customer = kata-kata bermakna di AKHIR remark (setelah noise)
        rest = re.sub(r'^(?:TRSF|BYR VIA)\s+E-BANKING\s+', '', desc, flags=re.I).strip()
        # BI-FAST: strip "BI-FAST BIF TRANSFER 009" atau "BI-FAST BIF BIFBI 008"
        rest = re.sub(r'^BI-FAST\s+\S+\s+(?:TRANSFER|BIFBI|DEBIT|CREDIT)\s+\d+\s*', '', rest, flags=re.I).strip()
        rest = re.sub(r'^BI-FAST\s+\S+\s*', '', rest, flags=re.I).strip()
        # Bersihkan kode tanggal BCA di awal: 0201/FTSCY/WS95051
        rest = re.sub(r'^\d{4}/[\w]+/[\w]+\s*', '', rest).strip()
        rest = re.sub(r'^/[\w]+/[\w]+\s*', '', rest).strip()

        words = rest.split()
        if not words:
            return None

        # Scan dari AKHIR: ambil kata-kata nama customer.
        # Aturan:
        #  - Kata mengandung digit                       → STOP (kode referensi)
        #  - Kata keterangan (BELI, NOTA, dll)           → STOP
        #  - Huruf kecil panjang >2 char                 → STOP (keterangan)
        #  - Title Case setelah dapat ≥1 kata ALL CAPS   → STOP ("Tepung" sebelum "SRI YULIATI")
        #  - Singkatan 1 huruf KAPITAL (P, S, T)         → VALID (suffix perusahaan)
        BULAN = {
            'JANUARI','FEBRUARI','MARET','APRIL','MEI','JUNI','JULI','AGUSTUS',
            'SEPTEMBER','OKTOBER','NOVEMBER','DESEMBER','DES','NOV','OKT','SEP',
            'AGU','JUL','JUN','APR','MAR','FEB','JAN','JANUARY','FEBRUARY',
            'MARCH','MAY','JUNE','JULY','AUGUST','SEPTEMBER','OCTOBER',
            'NOVEMBER','DECEMBER',
        }
        KET_STOP = {
            'TGL','NOTA','INV','INVOICE','BAYAR','BYR','LUNAS','PEMBAYARAN',
            'PEMBELIAN','BELANJA','BELI','DAN','DLL','BILL','REF','SALDO',
            'TAGIHAN','DP','PELUNASAN','ANGSURAN','TAG','SO','PO','TRANSFER','KE','DARI',
        }

        def _is_name_w_bca(w, n_caps_got):
            wu = w.upper()
            if re.search(r'\d', w): return False          # digit dalam kata = kode
            if re.match(r'^[\d,./\-\+&:]+$', w): return False
            if wu in KET_STOP: return False
            if wu in BULAN: return False
            if w == w.lower() and len(w) > 2: return False  # huruf kecil panjang = keterangan
            if len(w) == 1: return w.isupper()              # singkatan KAPITAL selalu valid
            if w == w.upper(): return True                   # ALL CAPS = nama
            # Title Case: valid HANYA sebelum dapat kata ALL CAPS
            # Setelah dapat ALL CAPS, Title Case = keterangan di depan nama
            if w[0].isupper() and w[1:].islower():
                return n_caps_got == 0
            return True

        name_words = []
        n_caps = 0
        for w in reversed(words):
            wu = w.upper()
            if re.search(r'\d', w): break             # digit = kode referensi, stop
            if re.match(r'^[\d,./\-\+&:]+$', wu): break
            if re.match(r'^REF:', wu): break
            if wu in OWN_W and not name_words: break
            if _is_name_w_bca(w, n_caps):
                name_words.insert(0, w)
                if w == w.upper() and len(w) > 1:
                    n_caps += 1
                if len(name_words) >= 7: break
            else:
                break
        if not name_words:
            candidates = [w for w in words[-5:]
                          if not re.match(r'^[\d,./\-]+$', w)
                          and len(w) > 2 and w.upper() not in OWN_W
                          and w.upper() not in KET_STOP]
            if candidates:
                name_words = candidates[-4:]
            else:
                return None

        name = ' '.join(name_words).strip()
        name = re.sub(r'^[-:,/\.\s]+', '', name).strip()
        name = re.sub(r'^\d{1,2}[A-Za-z]{2,}\s+', '', name).strip()

        # Buang nama rekening sendiri dari awal (gunakan OWN_W hardcoded)
        nm_w = name.split()
        j = 0
        while j < len(nm_w) and nm_w[j].upper() in OWN_W:
            j += 1
        if j > 0 and j < len(nm_w):
            name = ' '.join(nm_w[j:])

        if name and len(name) >= 2:
            return name.title()
        return None
    # ── FORMAT MANDIRI khusus: nama setelah "DARI" tanpa separator ───────────
    # "MCM InhouseTrf DARI NURUL KURNIAWATI bayar tepung Nurul99101"
    # Ambil nama setelah DARI, stop di: kata keterangan, kode 9910x, Transfer Fee, angka
    # Format: "... MCM InhouseTrf DARI NAMA [detik] keterangan..."
    # Tangkap nama: stop di kata keterangan, angka panjang, atau kode
    if 'DARI' in up and ' | ' not in desc:
        m = re.search(r'\bDARI\s+(.+?)(?:\s+(?:Transfer|Fee|Pelunasan|bayar|Byr|nota|kerupuk|marinasi|warkop|finna|ffc|bwi|toko|\d+\s+Jan|\d+\s+Feb|\d{4,}|9910)|$)', up, re.I)
        if m:
            name = m.group(1).strip()
            # Buang angka detik 1-2 digit di tengah nama (mis: "ANJAR 51 RAKHMA" → "ANJAR RAKHMA")
            name = re.sub(r'\b\d{1,2}\b', '', name).strip()
            # Buang kode embedded di akhir (Nurul99101 → Nurul)
            name = re.sub(r'\d{3,}.*$', '', name).strip()
            name = re.sub(r'\s{2,}', ' ', name).strip()
            if len(name) >= 3 and not re.match(r'^[\d\s]+$', name):
                return name.title()

    if ' | ' in desc:
        parts = desc.split(' | ')
        keterangan = parts[0].strip()
        name_part  = parts[-1].strip()

        # Bersihkan kode di awal nama: "3QO--NAMA" → "NAMA"
        name_clean = re.sub(r'^[A-Z0-9\-]+--', '', name_part)
        name_clean = re.sub(r'^[\d\-/]+\s*', '', name_clean).strip()

        # AutoCr-PL = sistem, bukan nama customer → ambil dari keterangan
        if name_clean.upper() in ('AUTOCR-PL', 'AUTOCR', 'PL', ''):
            keterangan_clean = re.sub(r'\b[A-Z]{4,}/[A-Z0-9]+/[A-Z0-9]+\b', '', keterangan)
            keterangan_clean = re.sub(r'\bBCA\d+\b', '', keterangan_clean).strip()
            keterangan_clean = re.sub(r'\s+', ' ', keterangan_clean).strip()
            return keterangan_clean.title() if keterangan_clean else 'Transfer BCA'

        # Nama yang mengandung tanda kurung = nama cabang/kota, bukan customer
        # e.g. "(BCA AMPERA", "BCA VETERAN )", ") JKT"
        if re.search(r'[()]', name_clean):
            # Ambil kata bermakna sebelum kurung, atau pakai keterangan awal
            keterangan_clean = keterangan.upper()
            if 'SETORAN TUNAI' in keterangan_clean:
                return 'Setoran Tunai'
            return keterangan.title()[:30]

        if len(name_clean) >= 3:
            return name_clean.title()

    # 1. NBMB / IBIZ / ATM — ambil nama antara prefix dan TO
    m = re.match(r'^(?:NBMB|IBIZ|ATM[A-Z]*)\s+(.+?)\s+TO\s+', up)
    if m:
        name = m.group(1).strip()
        # Hilangkan suffix kode rekening (angka panjang di akhir)
        name = re.sub(r'\s+\d{6,}.*$', '', name).strip()
        return name.title()

    # 2. BFST[nomor rekening][NAMA CUSTOMER]:[KODE]
    m = re.match(r'^BFST\d+(.+?)(?::[A-Z]{6,}|\s+ESB:)', up)
    if m:
        name = m.group(1).strip()
        # Hapus single char atau 2 char di akhir (sisa kode terpotong)
        name = re.sub(r'\s+[A-Z]{1,2}$', '', name).strip()
        return name.title()

    # 3. CBM_[kode]_DF_IP_[NAMA CUSTOMER] atau CBM_[kode]_DF_D_[NAMA]
    m = re.match(r'^CBM_[^_]+_DF_[^_]+_(.+?)(?:\s+API_|\s+APILOAN_|$)', up)
    if m:
        name = m.group(1).strip()
        # Hilangkan suffix P/PT di akhir
        name = re.sub(r'\s+P(?:T)?$', '', name).strip()
        return name.title()

    # 4. RTGS#[NAMA PENGIRIM] RTGS STP / ESB:
    m = re.match(r'^RTGS#(.+?)(?:\s+(?:PT\s+)?RTGS\s+STP|\s*#|\s+ESB:|\s+\d{10,}|$)', up)
    if m:
        name = m.group(1).strip()
        return name.title()

    # 5. IFT_TO_[NAMA] — pembayaran pelunasan
    m = re.search(r'IFT_TO_(.+?)(?:\s+ESB:|\s*$)', up)
    if m:
        # Ambil teks sebelum IFT_TO sebagai keterangan (misal: SO0124-099921 PELUNASAN)
        before = up.split('IFT_TO_')[0].strip()
        # Cari kata keterangan bermakna (bukan kode)
        keterangan = re.findall(r'[A-Z]{3,}', before)
        keterangan = [k for k in keterangan if k not in ('ESB','CMSX','FROM','IFT')]
        if keterangan:
            return ' '.join(keterangan[-2:]).title()  # ambil 2 kata terakhir
        return 'Pelunasan'

    # 6. TRA:[kode] BEFX — incoming wire transfer
    if re.match(r'^TRA:', up):
        return 'Wire Transfer'

    # 7. TUNAI / SETORAN TUNAI
    if up.startswith('TUNAI') or up.startswith('SETORAN TUNAI'):
        return 'Setoran Tunai'

    # 8. QRIS
    if up.startswith('QRIS') or 'BRIMCRDT' in up:
        return 'QRIS'

    # 9. FROM[nomor] TO[nomor] — transfer antar rekening
    if re.match(r'^FROM\d+', up):
        return 'Transfer Antar Rek'

    # 9b. DARI [nomrek] KE [nomrek] — transfer antar rekening sendiri
    if re.match(r'^DARI\s+\d+', up):
        return 'Transfer Antar Rek'

    # 9c. Hanya kode ESB tanpa info lain — transaksi internal bank
    if re.match(r'^(?:\d+\s*;\s*)?ESB:', up):
        return 'ESB/Transfer'

    # 10. KASDA / SP2D — pembayaran pemerintah
    if 'KASDA' in up or 'SP2D' in up or 'SPAN' in up:
        return 'Kasda / SP2D'

    # 10. Nama lain — ambil 2-3 kata pertama, buang kode/suffix umum
    # Hapus suffix tidak bermakna
    clean = re.sub(r'\s+(INVOICE|TGL|NO|ESB|FROM|TO|BY)\b.*', '', up).strip()
    clean = re.sub(r'[-_]\w+', ' ', clean).strip()  # hapus kode dengan dash/underscore
    words = re.findall(r'[A-Z][A-Z]+', clean)
    words = [w for w in words if w not in ('ESB','INDS','CMSX','FROM','TO','BY','PT','CV','UD','THE')
             and not re.match(r'^\d+$', w) and len(w) > 1]
    return ' '.join(words[:3]).title() if words else desc[:30].strip()


def _rtgs_is_own_or_generic(desc, company_name):
    """
    RTGS# Non penjualan jika:
    1. Sender mengandung nama rekening sendiri, ATAU
    2. Sender hanya berisi nama kota / kata sistem (bukan nama customer nyata)
    """
    import re as _re
    up = desc.upper()
    m = _re.match(r'RTGS#(.+?)(?:\s+(?:PT\s+)?RTGS\s+STP|\s*#|\s+ESB:|\s+\d{10,}|$)', up)
    if not m:
        return False
    sender = m.group(1).strip()

    # Cek nama rekening sendiri
    if _contains_own_name(sender, company_name):
        return True

    # Sender hanya nama kota / kata sistem = settlement bank, bukan customer
    CITY_SYSTEM = {
        'SURABAYA','JAKARTA','BANDUNG','SEMARANG','MEDAN','MAKASSAR','DENPASAR',
        'YOGYAKARTA','PALEMBANG','BALIKPAPAN','SAMARINDA','MALANG','SOLO',
        'PONTIANAK','PEKANBARU','BATAM','PADANG','MANADO','AMBON','KUPANG',
        'BRIRSS','STP','RTGS','PT',
    }
    sender_words = set(sender.split())
    if sender_words and sender_words.issubset(CITY_SYSTEM):
        return True

    return False

def _categorize(desc, company_name='', debet=0, kredit=0):
    """
    Filosofi AGRESIF — default kredit dari luar = Penjualan, koreksi manual jika salah.

    URUTAN:
      1. Debet saja → Non penjualan
      2. Kredit = 0 → Non penjualan
      3. Nama rekening sendiri di deskripsi → Non penjualan
      4. NON_PENJ_WHOLE (whole-word regex) → Non penjualan
      5. NON_PENJ_KW (substring) → Non penjualan
      6. RTGS# → Penjualan kecuali sender = nama rekening sendiri
      7. PENJ_KW eksplisit → Penjualan
      8. DEFAULT: semua kredit dari luar = Penjualan
    """
    import re as _re
    up = (desc or '').upper()

    # Rule 1: Debet murni → Non penjualan
    if debet > 0 and kredit == 0:
        return 'Non penjualan'

    # Rule 2: Tidak ada kredit → Non penjualan
    if kredit == 0:
        return 'Non penjualan'

    # Rule 3: IFT_TO_ — selalu Penjualan (pembayaran masuk ke rekening sendiri)
    if 'IFT_TO_' in up:
        return 'Penjualan'

    # Rule 4: Pola penerimaan — [NAMA PENGIRIM] TO [NAMA SENDIRI]
    # Harus dicek SEBELUM _contains_own_name karena nama sendiri ada di bagian "TO ..."
    _up_no_ift = up.replace('IFT_TO_', 'IFT_XX_')
    m_to = _re.search(r'\bTO\s+(.+?)(?:\s+ESB:|\s+FROM|\s*$)', _up_no_ift)
    if m_to:
        after_to  = m_to.group(1).strip()
        before_to = _up_no_ift[:m_to.start()].strip()
        if _contains_own_name(after_to, company_name) and \
           not _contains_own_name(before_to, company_name):
            return 'Penjualan'

    # Rule 4b: Format BCA — TRSF E-BANKING CR / KR OTOMATIS / BYR VIA E-BANKING
    # Transaksi ini SELALU kredit masuk dari customer.
    # Nama rekening sendiri muncul di remark hanya sebagai referensi invoice — bukan berarti internal.
    BCA_INBOUND = ('TRSF E-BANKING', 'KR OTOMATIS', 'BYR VIA E-BANKING', 'SWITCHING CR',
                   'SWITCHING DR')
    if kredit > 0 and any(up.startswith(p) for p in BCA_INBOUND):
        return 'Penjualan'

    # Rule 5: Mengandung nama rekening sendiri → Non penjualan
    # KECUALI: ada nama pengirim lain setelah "DARI" (Mandiri MCM InhouseTrf)
    # Contoh: "belanja Pangan Lestari MCM InhouseTrf DARI HARDHANI JUNIARTI" → Penjualan
    if _contains_own_name(desc, company_name):
        # Cek: apakah ada nama pengirim setelah DARI yang BUKAN nama sendiri?
        m_dari = _re.search(r"\bDARI\s+([A-Za-z][A-Za-z\s]+?)(?:\s+(?:Transfer|Fee|Pelunasan|bayar|Byr|belanja|nota|kerupuk|marinasi|warkop|pembayaran|pangan|lestari|\d{4,}|9910)|99\d{3}|\s*$)", up, _re.I)
        if m_dari:
            nama_pengirim = m_dari.group(1).strip()
            # Nama pengirim bukan nama sendiri → ini Penjualan dari customer
            if nama_pengirim and not _contains_own_name(nama_pengirim, company_name):
                return 'Penjualan'
        # Tidak ada nama pengirim lain → memang Non penjualan
        return 'Non penjualan'


    # Rule 6: Whole-word keywords
    for pattern in NON_PENJ_WHOLE:
        if _re.search(pattern, up):
            return 'Non penjualan'

    # Rule 7: Substring keywords non-penjualan
    for kw in NON_PENJ_KW:
        if kw.upper() in up:
            return 'Non penjualan'

    # Rule 8: RTGS# — cek sender
    if up.startswith('RTGS#'):
        if _rtgs_is_own_or_generic(desc, company_name):
            return 'Non penjualan'
        return 'Penjualan'

    # Rule 7: Keyword Penjualan eksplisit
    for kw in PENJ_KW:
        if kw.upper() in up:
            return 'Penjualan'

    # Rule 8: DEFAULT — semua kredit dari pihak luar = Penjualan
    return 'Penjualan'


# ── Buat Excel ─────────────────────────────────────────────────────────────────
def build_excel(all_transactions, meta, out_path):
    from openpyxl.worksheet.datavalidation import DataValidation

    by_month = defaultdict(list)
    for tx in all_transactions:
        by_month[tx['month']].append(tx)

    month_order = [f"{m} {y}" for y in range(2024, 2027)
                   for m in MONTHS_ID if f"{m} {y}" in by_month]

    n_tx = len(all_transactions)   # total baris di Edit Penjualan
    EP   = "'Edit Penjualan'"       # nama sheet untuk formula cross-ref

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Edit Penjualan  (MASTER — user edit di sini)
    # Kolom: A=No | B=Periode | C=Tanggal | D=Deskripsi | E=Debet | F=Kredit | G=Kategori
    # ═══════════════════════════════════════════════════════════════════════════
    ws_ep = wb.active
    ws_ep.title = "Edit Penjualan"

    # Title
    ws_ep.merge_cells('A1:G1')
    ws_ep['A1'] = f"✏️  EDIT KATEGORI PENJUALAN  —  {meta['companyName']}  —  {meta['accountNo']}"
    ws_ep['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=12)
    ws_ep['A1'].fill      = af(CLR["title_bg"])
    ws_ep['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_ep.row_dimensions[1].height = 26

    ws_ep.merge_cells('A2:G2')
    ws_ep['A2'] = "👆  Ubah kolom KATEGORI (G) untuk menandai transaksi sebagai Penjualan atau Non penjualan — Summary otomatis terupdate"
    ws_ep['A2'].font      = Font(name='Arial', italic=True, color="7F7F7F", size=9)
    ws_ep['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_ep.row_dimensions[2].height = 16

    hdrs_ep = ["No","Periode","Posting Date","Transaction Description","Debet (Rp)","Kredit (Rp)","Kategori","Customer ✏️"]
    ws_ep.append(hdrs_ep)
    style_hdr(ws_ep, 3)
    ws_ep.row_dimensions[3].height = 26

    # Dropdown validation
    dv = DataValidation(
        type="list",
        formula1='"Penjualan,Non penjualan"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Nilai tidak valid",
        error='Pilih: "Penjualan" atau "Non penjualan"'
    )
    ws_ep.add_data_validation(dv)

    # Isi data
    for i, tx in enumerate(all_transactions):
        r = i + 4   # baris 4 dst
        penj = tx['kategori'] == 'Penjualan'
        cust = tx.get('customer') or (_extract_customer_name(tx['desc']) if tx['kategori'] == 'Penjualan' else '')
        ws_ep.append([
            i + 1,
            tx['month'],
            tx['date'],
            tx['desc'],
            tx['debet']  if tx['debet']  else None,
            tx['kredit'] if tx['kredit'] else None,
            tx['kategori'],
            cust
        ])
        bg = CLR["penj_bg"] if penj else (CLR["alt_bg"] if i % 2 == 0 else "FFFFFF")
        for c in range(1, 8):
            cell = ws_ep.cell(row=r, column=c)
            cell.fill   = af(bg)
            cell.border = thin_border()
            cell.font   = reg()
        ws_ep.cell(row=r, column=1).alignment = Alignment(horizontal='center')
        ws_ep.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws_ep.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        e = ws_ep.cell(row=r, column=5)
        f = ws_ep.cell(row=r, column=6)
        e.number_format = NUM_FMT; e.alignment = Alignment(horizontal='right')
        f.number_format = NUM_FMT; f.alignment = Alignment(horizontal='right')
        if tx['debet']:  e.font = reg(color=CLR["debet_fg"])
        if tx['kredit']: f.font = reg(color=CLR["kredit_fg"])
        kat_cell = ws_ep.cell(row=r, column=7)
        kat_cell.alignment = Alignment(horizontal='center')
        if penj: kat_cell.font = reg(bold=True, color=CLR["kredit_fg"])
        dv.add(kat_cell)

    for col, w in zip("ABCDEFGH", [6, 14, 20, 62, 18, 18, 16, 36]):
        ws_ep.column_dimensions[col].width = w
    ws_ep.freeze_panes = 'A4'
    ws_ep.auto_filter.ref = f"A3:H{n_tx + 3}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Summary  (formula SUMIFS ke Edit Penjualan)
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Summary")

    ws.merge_cells('A1:H1')
    ws['A1'] = "REKAP REKENING KORAN BRI"
    ws['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=13)
    ws['A1'].fill      = af(CLR["title_bg"])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:H2')
    ws['A2'] = f"No. Rek: {meta['accountNo']}   |   {meta['companyName']}   |   Periode: {meta['period']}"
    ws['A2'].font      = Font(name='Arial', bold=True, color=CLR["sub_fg"], size=10)
    ws['A2'].fill      = af(CLR["sub_bg"])
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.append([])  # baris 3 kosong

    hdrs = ["Periode","Jumlah Transaksi","Total Debet (Rp)","Total Kredit (Rp)",
            "Total Penjualan (Rp)","Saldo Awal (Rp)","Saldo Akhir (Rp)","Status"]
    ws.append(hdrs)
    style_hdr(ws, 4)
    ws.row_dimensions[4].height = 30

    # Range Edit Penjualan untuk formula (baris 4..n_tx+3)
    ep_b = f"{EP}!$B$4:$B${n_tx+3}"   # kolom Periode
    ep_e = f"{EP}!$E$4:$E${n_tx+3}"   # kolom Debet
    ep_f = f"{EP}!$F$4:$F${n_tx+3}"   # kolom Kredit
    ep_g = f"{EP}!$G$4:$G${n_tx+3}"   # kolom Kategori

    for i, m in enumerate(month_order):
        txs_m = by_month[m]
        r     = 5 + i
        sa    = txs_m[0]['balance'] - txs_m[0]['kredit'] + txs_m[0]['debet'] if txs_m else 0
        se    = txs_m[-1]['balance'] if txs_m else 0

        # Formula SUMIFS ke Edit Penjualan — otomatis update saat user edit Kategori
        f_jumlah  = f'=COUNTIF({ep_b},A{r})'
        f_debet   = f'=SUMIF({ep_b},A{r},{ep_e})'
        f_kredit  = f'=SUMIF({ep_b},A{r},{ep_f})'
        f_penj    = f'=SUMIFS({ep_f},{ep_b},A{r},{ep_g},"Penjualan")'

        ws.append([m, f_jumlah, f_debet, f_kredit, f_penj, sa, se, "OK"])

        bg = CLR["alt_bg"] if i % 2 == 0 else "FFFFFF"
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.font = reg(); cell.fill = af(bg); cell.border = thin_border()
        for c in [3, 4, 5, 6, 7]:
            ws.cell(row=r, column=c).number_format = NUM_FMT
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='right')
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=8).font      = reg(bold=True, color=CLR["kredit_fg"])
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')

    # Baris TOTAL — formula SUM dari baris bulan
    tr       = 5 + len(month_order)
    r_start  = 5
    r_end    = tr - 1
    ws.append([
        "TOTAL",
        f"=SUM(B{r_start}:B{r_end})",
        f"=SUM(C{r_start}:C{r_end})",
        f"=SUM(D{r_start}:D{r_end})",
        f"=SUM(E{r_start}:E{r_end})",
        "", "", ""
    ])
    style_total(ws, tr)
    ws.cell(row=tr, column=2).alignment = Alignment(horizontal='center')
    for c in [3, 4, 5]:
        ws.cell(row=tr, column=c).number_format = NUM_FMT
        ws.cell(row=tr, column=c).alignment = Alignment(horizontal='right')

    # Ringkasan dari PDF
    tr += 2
    ws.cell(row=tr, column=1).value = "── RINGKASAN DARI PDF ──"
    ws.cell(row=tr, column=1).font  = reg(bold=True, color="2E75B6")
    # Label disesuaikan untuk rekening pinjaman
    _is_pinj = meta.get("is_pinjaman", False)
    for label, val in [
        ("Baki Debet Awal"   if _is_pinj else "Saldo Awal",    meta["opening"]),
        ("Total Mutasi Debet" if _is_pinj else "Total Debet",   meta["totalDebet"]),
        ("Total Mutasi Kredit"if _is_pinj else "Total Kredit",  meta["totalKredit"]),
        ("Baki Debet Akhir"  if _is_pinj else "Saldo Akhir",   meta["closing"])]:
        tr += 1
        ws.cell(row=tr, column=1).value = label
        ws.cell(row=tr, column=1).font  = reg()
        ws.cell(row=tr, column=3).value = val
        ws.cell(row=tr, column=3).number_format = NUM_FMT
        ws.cell(row=tr, column=3).alignment = Alignment(horizontal='right')

    for col, w in zip("ABCDEFGH", [16, 18, 22, 22, 22, 22, 22, 10]):
        ws.column_dimensions[col].width = w

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET per Bulan  (Kategori = INDEX ke Edit Penjualan)
    # ═══════════════════════════════════════════════════════════════════════════
    # Buat index global: nomor baris di Edit Penjualan untuk tiap transaksi
    tx_ep_row = {id(tx): (idx + 4) for idx, tx in enumerate(all_transactions)}

    for m in month_order:
        txs_m = by_month[m]
        safe  = m.replace(' ', '_')[:31]
        ws2   = wb.create_sheet(title=safe)

        ws2.merge_cells('A1:G1')
        ws2['A1'] = f"TRANSAKSI {m.upper()}  —  {meta['companyName']}  —  No. Rek: {meta['accountNo']}"
        ws2['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=11)
        ws2['A1'].fill      = af(CLR["title_bg"])
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 22

        bal_label = "Baki Debet (Rp)" if meta.get("is_pinjaman") else "Balance (Rp)"
        ws2.append(["No","Posting Date","Transaction Description",
                    "Debet (Rp)","Kredit (Rp)", bal_label,"Kategori"])
        style_hdr(ws2, 2)
        ws2.row_dimensions[2].height = 26

        td_tot = tk_tot = 0
        for i, tx in enumerate(txs_m):
            ep_row = tx_ep_row[id(tx)]
            r      = i + 3
            penj   = tx['kategori'] == 'Penjualan'

            ws2.append([
                i + 1, tx['date'], tx['desc'],
                tx['debet']  if tx['debet']  else None,
                tx['kredit'] if tx['kredit'] else None,
                tx['balance'],
                f"={EP}!G{ep_row}"   # ← referensi langsung ke Edit Penjualan
            ])

            bg = CLR["penj_bg"] if penj else (CLR["alt_bg"] if i % 2 == 0 else "FFFFFF")
            for c in range(1, 8):
                cell = ws2.cell(row=r, column=c)
                cell.fill = af(bg); cell.border = thin_border(); cell.font = reg()
            ws2.cell(row=r, column=1).alignment = Alignment(horizontal='center')
            ws2.cell(row=r, column=2).alignment = Alignment(horizontal='center')
            d = ws2.cell(row=r, column=4)
            k = ws2.cell(row=r, column=5)
            b = ws2.cell(row=r, column=6)
            d.number_format = NUM_FMT; d.alignment = Alignment(horizontal='right')
            k.number_format = NUM_FMT; k.alignment = Alignment(horizontal='right')
            b.number_format = NUM_FMT; b.alignment = Alignment(horizontal='right')
            if tx['debet']:  d.font = reg(color=CLR["debet_fg"])
            if tx['kredit']: k.font = reg(color=CLR["kredit_fg"])
            ws2.cell(row=r, column=7).alignment = Alignment(horizontal='center')
            if penj: ws2.cell(row=r, column=7).font = reg(bold=True, color=CLR["kredit_fg"])
            td_tot += tx['debet']; tk_tot += tx['kredit']

        tr2 = len(txs_m) + 3
        ws2.append(["", "TOTAL", "", td_tot, tk_tot, "", ""])
        style_total(ws2, tr2)
        ws2.cell(row=tr2, column=2).alignment = Alignment(horizontal='center')
        ws2.cell(row=tr2, column=4).number_format = NUM_FMT
        ws2.cell(row=tr2, column=4).alignment = Alignment(horizontal='right')
        ws2.cell(row=tr2, column=5).number_format = NUM_FMT
        ws2.cell(row=tr2, column=5).alignment = Alignment(horizontal='right')

        for col, w in zip("ABCDEFG", [6, 20, 58, 20, 20, 20, 16]):
            ws2.column_dimensions[col].width = w
        ws2.freeze_panes = 'A3'
        ws2.auto_filter.ref = f"A2:G{tr2}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET Penjualan Customer  (data statis — snapshot saat export)
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Penjualan Customer")
    ws3.merge_cells('A1:E1')
    ws3['A1'] = "PENJUALAN CUSTOMER"
    ws3['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=11)
    ws3['A1'].fill      = af(CLR["title_bg"])
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 22
    ws3.append(["Periode","Posting Date","Transaction Description","Kredit (Rp)","Customer"])
    style_hdr(ws3, 2)

    penj_txs = [tx for tx in all_transactions if tx['kategori'] == 'Penjualan']
    for i, tx in enumerate(penj_txs):
        ep_row = tx_ep_row[id(tx)]
        r      = i + 3
        ws3.append([
            tx['month'], tx['date'], tx['desc'],
            tx['kredit'],
            f"={EP}!H{ep_row}"   # Customer dari Edit Penjualan (bisa diedit user)
        ])
        bg = CLR["alt_bg"] if i % 2 == 0 else "FFFFFF"
        for c in range(1, 6):
            cell = ws3.cell(row=r, column=c)
            cell.fill = af(bg); cell.border = thin_border(); cell.font = reg()
        ws3.cell(row=r, column=4).number_format = NUM_FMT
        ws3.cell(row=r, column=4).alignment     = Alignment(horizontal='right')
        ws3.cell(row=r, column=4).font          = reg(color=CLR["kredit_fg"])

    for col, w in zip("ABCDE", [14, 20, 58, 20, 50]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = 'A3'

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET Customer Summary
    # Kolom B = Nama Customer (bisa diedit manual untuk menyesuaikan nama)
    # Kolom C = COUNTIFS → otomatis terhitung dari Edit Penjualan kolom H
    # Kolom D = SUMIFS   → otomatis terhitung dari Edit Penjualan kolom H
    # Sehingga: edit nama di EP!H → CS otomatis update
    #           edit nama di CS!B → bisa ubah "kunci" grouping
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Customer Summary")
    ws4.merge_cells('A1:D1')
    ws4['A1'] = f"RINGKASAN CUSTOMER PENJUALAN  —  {meta.get('period','')}"
    ws4['A1'].font      = Font(name='Arial', bold=True, color=CLR["title_fg"], size=11)
    ws4['A1'].fill      = af(CLR["title_bg"])
    ws4['A1'].alignment = Alignment(horizontal='center')
    ws4.row_dimensions[1].height = 22
    ws4.append(["No", "Customer ✏️", "Jml Transaksi", "Total Kredit (Rp)"])
    style_hdr(ws4, 2)
    # Tambah catatan di header
    ws4.cell(2, 2).comment = None  # clear existing

    # Range Edit Penjualan untuk formula
    ep_g_range = f"{EP}!$G$4:$G${n_tx+3}"   # kolom Kategori
    ep_h_range = f"{EP}!$H$4:$H${n_tx+3}"   # kolom Customer
    ep_f_range = f"{EP}!$F$4:$F${n_tx+3}"   # kolom Kredit

    # Kumpulkan daftar unik nama customer, urutkan dari terbesar kredit
    cs_map = {}
    for tx in penj_txs:
        cust_name = (tx.get('customer') or '').strip()
        if not cust_name:
            cust_name = _extract_customer_name(tx['desc'])
        if not cust_name:
            cust_name = '(Tidak Diketahui)'
        if cust_name not in cs_map:
            cs_map[cust_name] = {'customer': cust_name, 'kredit': 0}
        cs_map[cust_name]['kredit'] += tx['kredit']

    sorted_rows = sorted(cs_map.values(), key=lambda x: -x['kredit'])

    for i, rv in enumerate(sorted_rows):
        rn = i + 3
        bg = CLR["alt_bg"] if i % 2 == 0 else "FFFFFF"

        # Kolom A: No
        ws4.cell(rn, 1).value     = i + 1
        ws4.cell(rn, 1).alignment = Alignment(horizontal='center')

        # Kolom B: Nama customer (nilai — bisa diedit, jadi kunci formula)
        ws4.cell(rn, 2).value = rv['customer']

        # Kolom C: Jml Transaksi — COUNTIFS formula (otomatis update)
        # =COUNTIFS('Edit Penjualan'!$G$4:$G$N,"Penjualan",'Edit Penjualan'!$H$4:$H$N,B3)
        ws4.cell(rn, 3).value     = f'=COUNTIFS({ep_g_range},"Penjualan",{ep_h_range},B{rn})'
        ws4.cell(rn, 3).alignment = Alignment(horizontal='center')

        # Kolom D: Total Kredit — SUMIFS formula (otomatis update)
        # =SUMIFS('Edit Penjualan'!$F$4:$F$N,'Edit Penjualan'!$G$4:$G$N,"Penjualan",'Edit Penjualan'!$H$4:$H$N,B3)
        ws4.cell(rn, 4).value          = f'=SUMIFS({ep_f_range},{ep_g_range},"Penjualan",{ep_h_range},B{rn})'
        ws4.cell(rn, 4).number_format  = NUM_FMT
        ws4.cell(rn, 4).alignment      = Alignment(horizontal='right')
        ws4.cell(rn, 4).font           = reg(color=CLR["kredit_fg"])

        for c in range(1, 5):
            cell = ws4.cell(rn, c)
            cell.fill   = af(bg)
            cell.border = thin_border()
            if cell.font == reg():
                pass
            else:
                cell.font = reg()

    # Baris total — formula SUM dari kolom C dan D
    total_row = len(sorted_rows) + 3
    ws4.cell(total_row, 1).value = "TOTAL"
    if sorted_rows:
        ws4.cell(total_row, 3).value = f'=SUM(C3:C{total_row-1})'
        ws4.cell(total_row, 4).value = f'=SUM(D3:D{total_row-1})'
    for c in range(1, 5):
        cell = ws4.cell(total_row, c)
        cell.font   = Font(name='Arial', bold=True, size=10)
        cell.fill   = af(CLR["sub_bg"])
        cell.border = thin_border()
        if c == 4:
            cell.number_format = NUM_FMT
            cell.alignment     = Alignment(horizontal='right')
            cell.font          = Font(name='Arial', bold=True, color=CLR["kredit_fg"], size=10)
        if c in (1, 3):
            cell.alignment = Alignment(horizontal='center')

    # Tambah catatan panduan di bawah tabel
    note_row = total_row + 2
    ws4.cell(note_row, 1).value = "ℹ️ Cara edit nama customer:"
    ws4.cell(note_row, 1).font  = Font(name='Arial', italic=True, color="808080", size=9)
    ws4.merge_cells(f'A{note_row}:D{note_row}')
    note_row2 = total_row + 3
    ws4.cell(note_row2, 1).value = "   1. Edit kolom H (Customer ✏️) di sheet 'Edit Penjualan' → Jml & Total di sheet ini otomatis update"
    ws4.cell(note_row2, 1).font  = Font(name='Arial', italic=True, color="808080", size=9)
    ws4.merge_cells(f'A{note_row2}:D{note_row2}')
    note_row3 = total_row + 4
    ws4.cell(note_row3, 1).value = "   2. Edit kolom B (Customer) di sheet ini untuk menggabungkan/mengganti nama customer dalam ringkasan ini"
    ws4.cell(note_row3, 1).font  = Font(name='Arial', italic=True, color="808080", size=9)
    ws4.merge_cells(f'A{note_row3}:D{note_row3}')

    for col, w in zip("ABCD", [6, 40, 16, 22]):
        ws4.column_dimensions[col].width = w

    wb.save(out_path)


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("=" * 60)
        print("  BRI Rekening Koran → Excel Rekap")
        print("=" * 60)
        print("\nCara pakai:")
        print("  python bri_rekap.py file.pdf")
        print("  python bri_rekap.py folder/")
        sys.exit(0)

    input_path = Path(sys.argv[1])

    if input_path.is_dir():
        pdf_files = sorted(input_path.glob("*.pdf"))
        if not pdf_files:
            print(f"Tidak ada file PDF di folder: {input_path}"); sys.exit(1)
        print(f"Ditemukan {len(pdf_files)} file PDF")
    elif input_path.is_file():
        pdf_files = [input_path]
    else:
        print(f"File/folder tidak ditemukan: {input_path}"); sys.exit(1)

    all_transactions = []
    meta_combined    = {"accountNo":"","companyName":"","period":"",
                        "opening":0,"totalDebet":0,"totalKredit":0,"closing":0}
    periods = []

    for pdf_file in pdf_files:
        print(f"\nMemproses: {pdf_file.name}")
        meta, txs = parse_pdf(pdf_file)
        print(f"  Rekening  : {meta['accountNo']} — {meta['companyName']}")
        print(f"  Periode   : {meta['period']}")
        print(f"  Transaksi : {len(txs)}")
        all_transactions.extend(txs)
        if not meta_combined["accountNo"] and meta["accountNo"]:
            meta_combined["accountNo"]   = meta["accountNo"]
            meta_combined["companyName"] = meta["companyName"]
        meta_combined["totalDebet"]  += meta["totalDebet"]
        meta_combined["totalKredit"] += meta["totalKredit"]
        if meta["period"]: periods.append(meta["period"])
        if not meta_combined["opening"] and meta["opening"]:
            meta_combined["opening"] = meta["opening"]
        if meta["closing"]:
            meta_combined["closing"] = meta["closing"]

    if periods:
        meta_combined["period"] = f"{periods[0].split(' - ')[0]} - {periods[-1].split(' - ')[-1]}"

    acc = meta_combined["accountNo"] or "rekening"
    if input_path.is_file():
        out = Path.cwd() / f"rekap_{input_path.stem}.xlsx"
    else:
        out = Path.cwd() / f"rekap_bri_{acc}.xlsx"

    print(f"\nMembuat Excel...")
    build_excel(all_transactions, meta_combined, str(out))

    print(f"\n✓ SELESAI! File tersimpan di:")
    print(f"  {out.resolve()}")
    print(f"\n  Total transaksi  : {len(all_transactions)}")
    penj = [t for t in all_transactions if t['kategori'] == 'Penjualan']
    print(f"  Penjualan        : {len(penj)} transaksi")
    months = sorted(set(t['month'] for t in all_transactions))
    print(f"  Periode          : {', '.join(months)}")


if __name__ == "__main__":
    main()
